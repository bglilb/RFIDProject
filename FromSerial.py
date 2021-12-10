# FromSerial.py
# Reads data from serial line and writes to an excel sheet

import openpyxl
import time
import serial

def inRoster(uuidSum):
    for i in range(numStudents):
        if uuidSum == keys[i]:
            return i
    
    return -1

#File Path to Attendance Sheet
path = "//Users//BLGlilB//Documents//EE193//Final Project//Attendance.xlsx"

# Setting up workbook for roster sheet
names = []
keys = []
wb = openpyxl.load_workbook(path)
wb.active = 0
roster = wb.active
numStudents = roster.max_row

for i in range(1,numStudents+1):
    names.append(roster.cell(row = i, column = 1).value) #Adding Names to List
    keysum = 0
    for j in range(2,6):
        currcell= roster.cell(row = i, column = j).value
        cellInt = int(currcell,16) #Convert String to int
        #hexID = hex(cellInt)
        keysum += cellInt

    keys.append(keysum) #Adding Keys to List

# Setting up workbook for attendance sheet
wb.active = 1
attend = wb.active #Second Sheet is Attendance

caughtUp = False
numCheckIn = 0
# Allows for Attendance to be added to the end of the Excel Sheet
while not caughtUp:
    if attend.cell(numCheckIn + 2, 1).value is None:
        caughtUp = True
    else: 
        numCheckIn += 1


# Setting up serial connection
ser = serial.Serial('/dev/cu.usbmodem14101',9600) #CHANGE TO CORRECT SERIAL PORT

running = False

print('Welcome to Class!')
command = input('Would you like to start taking attendance? (y/n): ')
if command == 'y':
    running = True
    print('Hit Ctrl + C to End Program')
    print('Welcome Students!')

try:
    while running:
        # Reading data from port
        uuidSum = 0
        for i in range(4):
            b = ser.readline() #Read In String
            string_n = b.decode() #Decode Binary
            string = string_n.rstrip() #Removes \r and \n
            intID = int(string) #Convert String from Serial to Int
            uuidSum += intID #Creates a Single Sum for ID Key

        student = inRoster(uuidSum) #Checks if student in roster and what student num
        if (student != -1):
            print('Welcome ' + names[student] + '!')
            dateInfo = time.localtime(time.time())
            date = str(dateInfo.tm_mon) + "/" + str(dateInfo.tm_mday) + "/" + str(dateInfo.tm_year)
            timeInfo = str(dateInfo.tm_hour) + ":" + str(dateInfo.tm_min) + ":" + str(dateInfo.tm_sec)
            attend.cell(numCheckIn + 2, 1).value = names[student] #Col 1 is name
            attend.cell(numCheckIn + 2, 2).value = "Here!" #Col 2 is here
            attend.cell(numCheckIn + 2, 3).value = date #Col 3 is date
            attend.cell(numCheckIn + 2, 4).value = timeInfo #Col 1 is time
            numCheckIn += 1

        wb.save("Attendance.xlsx")
except:
    print('\nThank You! Come Againg Soon :)')
    wb.save("Attendance.xlsx")
